import os
import re
from io import BytesIO
from datetime import date, datetime
from flask import Flask, render_template, request, redirect, url_for, send_file, jsonify, abort, flash
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
database_url = os.environ.get("DATABASE_URL") or "sqlite:///rsvp.db"
if database_url.startswith("postgres://"):
    database_url = database_url.replace("postgres://", "postgresql://", 1)
app.config["SQLALCHEMY_DATABASE_URI"] = database_url
if os.environ.get("DATABASE_URL"):
    app.config["SECRET_KEY"] = os.environ["SECRET_KEY"]
else:
    app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-only-change-me")
db = SQLAlchemy(app)

login_manager = LoginManager(app)
login_manager.login_view = "login"


# ── Constants ─────────────────────────────────────────────────────────────────

EVENT_TYPES = ["Dinner", "Party", "Weekend", "Hunt", "Corporate", "Other"]
CHANNELS = ["WhatsApp", "Email", "Call", "Live", "SMS", "Other"]


# ── Models ────────────────────────────────────────────────────────────────────

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(200), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    events = db.relationship("Event", backref="user", cascade="all, delete-orphan")
    guests = db.relationship("Guest", backref="user", cascade="all, delete-orphan")


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


class Event(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    name = db.Column(db.String(200), nullable=False)
    event_type = db.Column(db.String(50), nullable=False, default="Other")
    location = db.Column(db.String(200), nullable=True, default="")
    date = db.Column(db.Date, nullable=False)
    date_created = db.Column(db.Date, nullable=True)
    date_edited = db.Column(db.DateTime, nullable=True)
    notes = db.Column(db.Text, default="")
    target_attendees = db.Column(db.Integer, nullable=True)
    invitations = db.relationship("Invitation", backref="event", cascade="all, delete-orphan")


class Guest(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)
    first_name = db.Column(db.String(100), nullable=False)
    last_name = db.Column(db.String(100), nullable=True, default="")
    gender = db.Column(db.String(10), nullable=False)
    is_me = db.Column(db.Boolean, default=False)
    notes = db.Column(db.Text, default="")
    date_created = db.Column(db.DateTime, nullable=True)
    date_edited = db.Column(db.DateTime, nullable=True)
    invitations = db.relationship("Invitation", backref="guest", cascade="all, delete-orphan")

    @property
    def full_name(self):
        if self.last_name:
            return f"{self.first_name} {self.last_name}"
        return self.first_name


class Invitation(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    event_id = db.Column(db.Integer, db.ForeignKey("event.id"), nullable=False)
    guest_id = db.Column(db.Integer, db.ForeignKey("guest.id"), nullable=False)
    status = db.Column(db.String(20), nullable=False, default="Not Sent")
    channel = db.Column(db.String(50), nullable=True, default="")
    date_invited = db.Column(db.Date, nullable=True)
    date_responded = db.Column(db.Date, nullable=True)
    notes = db.Column(db.Text, nullable=True, default="")


# ── Auth routes ───────────────────────────────────────────────────────────────

@app.route("/signup", methods=["GET", "POST"])
def signup():
    if current_user.is_authenticated:
        return redirect(url_for("home"))
    if request.method == "POST":
        email = request.form["email"].strip().lower()
        password = request.form["password"]
        if not email or "@" not in email:
            return render_template("signup.html", error="Valid email is required")
        if User.query.filter_by(email=email).first():
            return render_template("signup.html", error="Email already registered")
        if len(password) < 6:
            return render_template("signup.html", error="Password must be at least 6 characters")
        user = User(email=email, password_hash=generate_password_hash(password))
        db.session.add(user)
        db.session.commit()
        login_user(user)
        return redirect(url_for("home"))
    return render_template("signup.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("home"))
    if request.method == "POST":
        email = request.form["email"].strip().lower()
        password = request.form["password"]
        user = User.query.filter_by(email=email).first()
        if not user or not check_password_hash(user.password_hash, password):
            return render_template("login.html", error="Invalid email or password")
        login_user(user)
        next_page = request.args.get("next")
        if next_page and (not next_page.startswith("/") or next_page.startswith("//")):
            next_page = None
        return redirect(next_page or url_for("home"))
    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))


# ── Settings ──────────────────────────────────────────────────────────────────

@app.route("/settings")
@login_required
def settings():
    return render_template("settings.html")


@app.route("/settings/load-sample-data", methods=["POST"])
@login_required
def load_sample_data():
    seed(current_user.id)
    flash("Sample data loaded successfully.")
    return redirect(url_for("settings"))


@app.route("/settings/reset-sample-data", methods=["POST"])
@login_required
def reset_sample_data():
    uid = current_user.id
    Invitation.query.filter(Invitation.event.has(user_id=uid)).delete(synchronize_session=False)
    Event.query.filter_by(user_id=uid).delete()
    Guest.query.filter_by(user_id=uid).delete()
    db.session.commit()
    seed(uid)
    flash("All data reset to sample data.")
    return redirect(url_for("settings"))


# ── Event routes ──────────────────────────────────────────────────────────────

@app.route("/")
@login_required
def home():
    events = Event.query.filter_by(user_id=current_user.id).order_by(Event.date).all()
    me_exists = Guest.query.filter_by(user_id=current_user.id, is_me=True).first() is not None
    return render_template("home.html", events=events, event_types=EVENT_TYPES, today_date=date.today(), me_exists=me_exists)


@app.route("/event/add", methods=["POST"])
@login_required
def add_event():
    try:
        event_date = date.fromisoformat(request.form["date"])
    except (ValueError, KeyError):
        return "Invalid date", 400
    event = Event(
        user_id=current_user.id,
        name=request.form["name"],
        event_type=request.form["event_type"],
        location=request.form.get("location", ""),
        date=event_date,
        date_created=date.today(),
        notes=request.form.get("notes", ""),
        target_attendees=request.form.get("target_attendees", type=int) or None,
    )
    db.session.add(event)
    db.session.commit()
    if request.form.get("include_me"):
        me = Guest.query.filter_by(user_id=current_user.id, is_me=True).first()
        if me:
            inv = Invitation(event_id=event.id, guest_id=me.id,
                             status="Attending", date_invited=date.today(),
                             date_responded=date.today())
            db.session.add(inv)
            db.session.commit()
    return redirect(url_for("event_detail", event_id=event.id))


@app.route("/event/<int:event_id>")
@login_required
def event_detail(event_id):
    event = Event.query.get_or_404(event_id)
    if event.user_id != current_user.id:
        return redirect(url_for("home"))
    return render_template("event_detail.html", event=event, channels=CHANNELS, event_types=EVENT_TYPES)


@app.route("/event/<int:event_id>/edit", methods=["POST"])
@login_required
def edit_event(event_id):
    event = Event.query.get_or_404(event_id)
    if event.user_id != current_user.id:
        return redirect(url_for("home"))
    try:
        event_date = date.fromisoformat(request.form["date"])
    except (ValueError, KeyError):
        return "Invalid date", 400
    event.name = request.form["name"]
    event.event_type = request.form["event_type"]
    event.location = request.form.get("location", "")
    event.date = event_date
    event.notes = request.form.get("notes", "")
    event.target_attendees = request.form.get("target_attendees", type=int) or None
    event.date_edited = datetime.now()
    db.session.commit()
    return redirect(url_for("event_detail", event_id=event.id))


@app.route("/event/<int:event_id>/delete", methods=["POST"])
@login_required
def delete_event(event_id):
    event = Event.query.get_or_404(event_id)
    if event.user_id != current_user.id:
        return redirect(url_for("home"))
    db.session.delete(event)
    db.session.commit()
    return redirect(url_for("home"))


# ── Guest routes ──────────────────────────────────────────────────────────────

@app.route("/guests")
@login_required
def guests():
    all_guests = Guest.query.filter_by(user_id=current_user.id).order_by(Guest.is_me.desc(), Guest.last_name, Guest.first_name).all()
    return render_template("guests.html", guests=all_guests)


@app.route("/guest/add", methods=["POST"])
@login_required
def add_guest():
    is_me = bool(request.form.get("is_me"))
    if is_me:
        Guest.query.filter_by(user_id=current_user.id, is_me=True).update({"is_me": False})
    guest = Guest(
        user_id=current_user.id,
        first_name=request.form["first_name"],
        last_name=request.form.get("last_name", ""),
        gender=request.form["gender"],
        is_me=is_me,
        notes=request.form.get("notes", ""),
        date_created=datetime.now(),
    )
    db.session.add(guest)
    db.session.commit()
    return redirect(url_for("guests"))


@app.route("/guest/<int:guest_id>/edit", methods=["POST"])
@login_required
def edit_guest(guest_id):
    guest = Guest.query.get_or_404(guest_id)
    if guest.user_id != current_user.id:
        return redirect(url_for("guests"))
    is_me = bool(request.form.get("is_me"))
    if is_me and not guest.is_me:
        Guest.query.filter_by(user_id=current_user.id, is_me=True).update({"is_me": False})
    guest.first_name = request.form["first_name"]
    guest.last_name = request.form.get("last_name", "")
    guest.gender = request.form["gender"]
    guest.is_me = is_me
    guest.notes = request.form.get("notes", "")
    guest.date_edited = datetime.now()
    db.session.commit()
    return redirect(url_for("guests"))


@app.route("/guest/<int:guest_id>/delete", methods=["POST"])
@login_required
def delete_guest(guest_id):
    guest = Guest.query.get_or_404(guest_id)
    if guest.user_id != current_user.id:
        return redirect(url_for("guests"))
    db.session.delete(guest)
    db.session.commit()
    return redirect(url_for("guests"))


# ── Invitation routes ─────────────────────────────────────────────────────────

@app.route("/invitation/<int:invitation_id>/send", methods=["POST"])
@login_required
def toggle_send_invitation(invitation_id):
    invitation = Invitation.query.get_or_404(invitation_id)
    if invitation.event.user_id != current_user.id:
        abort(403)
    if invitation.status == "Not Sent":
        invitation.status = "Pending"
        invitation.date_invited = date.today()
    else:
        invitation.status = "Not Sent"
        invitation.date_invited = None
        invitation.date_responded = None
    invitation.event.date_edited = datetime.now()
    db.session.commit()
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify(
            status=invitation.status,
            date_invited=invitation.date_invited.strftime("%b %d, %Y") if invitation.date_invited else "",
            date_invited_iso=invitation.date_invited.isoformat() if invitation.date_invited else "",
        )
    return redirect(url_for("event_detail", event_id=invitation.event_id))


@app.route("/invitation/<int:invitation_id>/update", methods=["POST"])
@login_required
def update_invitation(invitation_id):
    invitation = Invitation.query.get_or_404(invitation_id)
    if invitation.event.user_id != current_user.id:
        abort(403)
    new_status = request.form.get("status")
    if not new_status:
        return jsonify(error="Missing status"), 400
    if new_status != invitation.status:
        invitation.status = new_status
        if new_status in ("Attending", "Declined"):
            invitation.date_responded = date.today()
        elif new_status == "Pending":
            invitation.date_responded = None
    invitation.event.date_edited = datetime.now()
    db.session.commit()
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify(
            status=invitation.status,
            date_responded=invitation.date_responded.strftime("%b %d, %Y") if invitation.date_responded else "",
            date_responded_iso=invitation.date_responded.isoformat() if invitation.date_responded else "",
        )
    return redirect(url_for("event_detail", event_id=invitation.event_id))


@app.route("/invitation/<int:invitation_id>/delete", methods=["POST"])
@login_required
def remove_invitation(invitation_id):
    invitation = Invitation.query.get_or_404(invitation_id)
    if invitation.event.user_id != current_user.id:
        abort(403)
    event_id = invitation.event_id
    invitation.event.date_edited = datetime.now()
    db.session.delete(invitation)
    db.session.commit()
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify(ok=True)
    return redirect(url_for("event_detail", event_id=event_id))


# ── API endpoints ─────────────────────────────────────────────────────────────

@app.route("/api/guests/bulk-create", methods=["POST"])
@login_required
def bulk_create_guests():
    data = request.get_json()
    guests_data = data.get("guests", [])
    added = []
    for g_data in guests_data:
        first_name = g_data.get("first_name", "").strip()
        if not first_name:
            continue
        guest = Guest(
            user_id=current_user.id,
            first_name=first_name,
            last_name=g_data.get("last_name", "").strip(),
            gender=g_data.get("gender", "Male"),
            notes=g_data.get("notes", "").strip(),
            date_created=datetime.now()
        )
        db.session.add(guest)
        db.session.flush()
        added.append({
            "id": guest.id, "first_name": guest.first_name,
            "last_name": guest.last_name or "", "gender": guest.gender,
            "notes": guest.notes or "", "is_me": False,
            "date_created": guest.date_created.isoformat()
        })
    db.session.commit()
    return jsonify(added=added)


@app.route("/api/guest/<int:guest_id>/gender", methods=["POST"])
@login_required
def update_guest_gender(guest_id):
    guest = Guest.query.get_or_404(guest_id)
    if guest.user_id != current_user.id:
        abort(403)
    data = request.get_json()
    guest.gender = data.get("gender", guest.gender)
    guest.date_edited = datetime.now()
    db.session.commit()
    return jsonify(ok=True)


@app.route("/api/guest/<int:guest_id>/name", methods=["POST"])
@login_required
def update_guest_name(guest_id):
    guest = Guest.query.get_or_404(guest_id)
    if guest.user_id != current_user.id:
        abort(403)
    data = request.get_json()
    guest.first_name = data.get("first_name", guest.first_name)
    guest.last_name = data.get("last_name", guest.last_name or "")
    guest.date_edited = datetime.now()
    db.session.commit()
    return jsonify(ok=True, full_name=guest.full_name)


@app.route("/api/guest/<int:guest_id>/notes", methods=["POST"])
@login_required
def update_guest_notes(guest_id):
    guest = Guest.query.get_or_404(guest_id)
    if guest.user_id != current_user.id:
        abort(403)
    data = request.get_json()
    guest.notes = data.get("notes", "")
    guest.date_edited = datetime.now()
    db.session.commit()
    return jsonify(ok=True)


@app.route("/api/guest/<int:guest_id>/is-me", methods=["POST"])
@login_required
def update_guest_is_me(guest_id):
    guest = Guest.query.get_or_404(guest_id)
    if guest.user_id != current_user.id:
        abort(403)
    data = request.get_json()
    is_me = data.get("is_me", False)
    if is_me and not guest.is_me:
        Guest.query.filter_by(user_id=current_user.id, is_me=True).update({"is_me": False})
    guest.is_me = is_me
    guest.date_edited = datetime.now()
    db.session.commit()
    return jsonify(ok=True, is_me=guest.is_me)


@app.route("/api/invitation/<int:invitation_id>/field", methods=["POST"])
@login_required
def update_invitation_field(invitation_id):
    invitation = Invitation.query.get_or_404(invitation_id)
    if invitation.event.user_id != current_user.id:
        abort(403)
    data = request.get_json()
    field = data.get("field")
    value = data.get("value", "")
    if field == "channel":
        invitation.channel = value
    elif field == "notes":
        invitation.notes = value
    else:
        return jsonify(error="Invalid field"), 400
    invitation.event.date_edited = datetime.now()
    db.session.commit()
    return jsonify(ok=True)


@app.route("/api/event/<int:event_id>/notes", methods=["POST"])
@login_required
def update_event_notes(event_id):
    event = Event.query.get_or_404(event_id)
    if event.user_id != current_user.id:
        abort(403)
    data = request.get_json()
    event.notes = data.get("notes", "")
    event.date_edited = datetime.now()
    db.session.commit()
    return jsonify(ok=True)


@app.route("/api/event/<int:event_id>/available-guests")
@login_required
def api_available_guests(event_id):
    event = Event.query.get_or_404(event_id)
    if event.user_id != current_user.id:
        abort(403)
    invited_ids = {inv.guest_id for inv in event.invitations}
    all_guests = Guest.query.filter_by(user_id=current_user.id).order_by(Guest.first_name, Guest.last_name).all()
    result = []
    for g in all_guests:
        result.append({
            "id": g.id, "first_name": g.first_name, "last_name": g.last_name or "",
            "gender": g.gender, "already_invited": g.id in invited_ids
        })
    return jsonify(guests=result)


@app.route("/api/event/<int:event_id>/bulk-add", methods=["POST"])
@login_required
def bulk_add_guests(event_id):
    event = Event.query.get_or_404(event_id)
    if event.user_id != current_user.id:
        abort(403)
    data = request.get_json()
    guest_ids = data.get("guest_ids", [])
    invited_ids = {inv.guest_id for inv in event.invitations}
    added = []
    for gid in guest_ids:
        if gid in invited_ids:
            continue
        guest = Guest.query.get(gid)
        if not guest or guest.user_id != current_user.id:
            continue
        inv = Invitation(event_id=event_id, guest_id=gid, status="Not Sent")
        db.session.add(inv)
        db.session.flush()
        added.append({
            "invitation_id": inv.id, "guest_id": guest.id,
            "first_name": guest.first_name, "last_name": guest.last_name or "",
            "gender": guest.gender, "status": "Not Sent"
        })
    if added:
        event.date_edited = datetime.now()
    db.session.commit()
    return jsonify(added=added)


@app.route("/api/event/<int:event_id>/bulk-create-and-invite", methods=["POST"])
@login_required
def bulk_create_and_invite(event_id):
    event = Event.query.get_or_404(event_id)
    if event.user_id != current_user.id:
        abort(403)
    data = request.get_json()
    guests_data = data.get("guests", [])
    added = []
    for g_data in guests_data:
        first_name = g_data.get("first_name", "").strip()
        if not first_name:
            continue
        guest = Guest(
            user_id=current_user.id,
            first_name=first_name,
            last_name=g_data.get("last_name", "").strip(),
            gender=g_data.get("gender", "Male"),
            notes=g_data.get("notes", "").strip(),
            date_created=datetime.now()
        )
        db.session.add(guest)
        db.session.flush()
        inv = Invitation(event_id=event_id, guest_id=guest.id, status="Not Sent")
        db.session.add(inv)
        db.session.flush()
        added.append({
            "invitation_id": inv.id, "guest_id": guest.id,
            "first_name": guest.first_name, "last_name": guest.last_name or "",
            "gender": guest.gender, "status": "Not Sent",
            "channel": "", "notes": "",
            "date_invited": "", "date_invited_iso": "",
            "date_responded": "", "date_responded_iso": ""
        })
    if added:
        event.date_edited = datetime.now()
    db.session.commit()
    return jsonify(added=added)


# ── Export helpers ─────────────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")


def _styled_sheet(wb, title, headers):
    ws = wb.active
    ws.title = re.sub(r"[/\\*?\[\]:]", "_", title)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    return ws


def _to_download(wb, filename):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name=filename, as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── Export routes ─────────────────────────────────────────────────────────────

@app.route("/export/events")
@login_required
def export_events():
    wb = Workbook()
    ws = _styled_sheet(wb, "Events", ["Name", "Type", "Date", "Location", "Invited", "Attending", "Notes"])
    for e in Event.query.filter_by(user_id=current_user.id).order_by(Event.date).all():
        attending = sum(1 for inv in e.invitations if inv.status == "Attending")
        ws.append([e.name, e.event_type, e.date.strftime("%Y-%m-%d"), e.location or "",
                   len(e.invitations), attending, e.notes or ""])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    return _to_download(wb, "events.xlsx")


@app.route("/export/guests")
@login_required
def export_guests():
    wb = Workbook()
    ws = _styled_sheet(wb, "Guests", ["Last Name", "First Name", "Gender", "Notes"])
    for g in Guest.query.filter_by(user_id=current_user.id).order_by(Guest.last_name, Guest.first_name).all():
        ws.append([g.last_name or "", g.first_name, g.gender, g.notes or ""])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    return _to_download(wb, "guests.xlsx")


@app.route("/export/event/<int:event_id>")
@login_required
def export_event_guests(event_id):
    event = Event.query.get_or_404(event_id)
    if event.user_id != current_user.id:
        abort(403)
    wb = Workbook()
    ws = _styled_sheet(wb, event.name[:31],
                       ["Last Name", "First Name", "Gender", "Sent", "Channel",
                        "Invited On", "Status", "Responded On", "Inv. Notes", "Guest Notes"])
    for inv in event.invitations:
        g = inv.guest
        sent = "Yes" if inv.status != "Not Sent" else "No"
        ws.append([g.last_name or "", g.first_name, g.gender, sent, inv.channel or "",
                   inv.date_invited.strftime("%Y-%m-%d") if inv.date_invited else "",
                   inv.status,
                   inv.date_responded.strftime("%Y-%m-%d") if inv.date_responded else "",
                   inv.notes or "", g.notes or ""])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    safe_name = re.sub(r"[^\w\-]", "_", event.name).strip("_").lower()
    return _to_download(wb, f"{safe_name}_guests.xlsx")


# ── Seed data ─────────────────────────────────────────────────────────────────

def seed(user_id):
    """Populate a user's account with sample data."""
    today = date.today()
    events = [
        # Past events
        Event(user_id=user_id, name="New Year's Eve Bash", event_type="Party", location="Rooftop Lounge, Miami", date=date(2025, 12, 31), date_created=date(2025, 11, 15), notes="Great turnout!", target_attendees=20),
        Event(user_id=user_id, name="Q4 Board Meeting", event_type="Corporate", location="HQ Boardroom, Chicago", date=date(2026, 1, 15), date_created=date(2025, 12, 1), notes="Quarterly review completed"),
        Event(user_id=user_id, name="Valentine's Dinner", event_type="Dinner", location="Le Petit Bistro, Paris", date=date(2026, 2, 14), date_created=date(2026, 1, 10), notes="Intimate dinner for couples", target_attendees=8),
        # Upcoming events
        Event(user_id=user_id, name="Annual Gala Dinner", event_type="Dinner", location="Grand Hotel, New York", date=date(2026, 4, 12), date_created=today, notes="Black tie event", target_attendees=10),
        Event(user_id=user_id, name="Team Building Retreat", event_type="Corporate", location="Lakehouse Resort, Vermont", date=date(2026, 5, 20), date_created=today, notes="Outdoor activities planned", target_attendees=6),
        Event(user_id=user_id, name="Summer Garden Party", event_type="Party", location="Riverside Park, Boston", date=date(2026, 7, 4), date_created=today, notes="Casual dress code"),
        Event(user_id=user_id, name="Lake House Weekend", event_type="Weekend", location="Lakehouse Resort, Vermont", date=date(2026, 8, 22), date_created=today, notes="Arrive Friday evening, depart Sunday", target_attendees=15),
        Event(user_id=user_id, name="Fall Pheasant Hunt", event_type="Hunt", location="Highland Estate, Montana", date=date(2026, 9, 15), date_created=today, notes="Bring warm layers", target_attendees=4),
    ]
    db.session.add_all(events)
    db.session.flush()

    now = datetime.now()
    guests = [
        Guest(user_id=user_id, first_name="Alice", last_name="Martin", gender="Female", notes="Vegetarian", date_created=now),
        Guest(user_id=user_id, first_name="James", last_name="Wilson", gender="Male", date_created=now),
        Guest(user_id=user_id, first_name="Sofia", last_name="Garcia", gender="Female", notes="Plus one confirmed", date_created=now),
        Guest(user_id=user_id, first_name="Oliver", last_name="Brown", gender="Male", date_created=now),
        Guest(user_id=user_id, first_name="Emma", last_name="Taylor", gender="Female", date_created=now),
        Guest(user_id=user_id, first_name="William", last_name="Anderson", gender="Male", notes="Allergic to nuts", date_created=now),
        Guest(user_id=user_id, first_name="Charlotte", last_name="Thomas", gender="Female", date_created=now),
        Guest(user_id=user_id, first_name="Henry", last_name="Jackson", gender="Male", date_created=now),
        Guest(user_id=user_id, first_name="Amelia", last_name="White", gender="Female", notes="Wheelchair accessible seating", date_created=now),
        Guest(user_id=user_id, first_name="Lucas", last_name="Harris", gender="Male", date_created=now),
        Guest(user_id=user_id, first_name="Isabella", last_name="Clark", gender="Female", date_created=now),
        Guest(user_id=user_id, first_name="Benjamin", last_name="Lee", gender="Male", notes="VIP", date_created=now),
        Guest(user_id=user_id, first_name="Mia", last_name="Robinson", gender="Female", date_created=now),
        Guest(user_id=user_id, first_name="Ethan", last_name="Wright", gender="Male", notes="Prefers window seat", date_created=now),
        Guest(user_id=user_id, first_name="Ava", last_name="Lopez", gender="Female", notes="Gluten-free", date_created=now),
        Guest(user_id=user_id, first_name="Noah", last_name="Mitchell", gender="Male", date_created=now),
        Guest(user_id=user_id, first_name="Lily", last_name="Perez", gender="Female", date_created=now),
        Guest(user_id=user_id, first_name="Daniel", last_name="Roberts", gender="Male", notes="Board member", date_created=now),
        Guest(user_id=user_id, first_name="Grace", last_name="Turner", gender="Female", date_created=now),
        Guest(user_id=user_id, first_name="Jack", last_name="Phillips", gender="Male", date_created=now),
        Guest(user_id=user_id, first_name="Chloe", last_name="Campbell", gender="Female", notes="Bringing plus one", date_created=now),
        Guest(user_id=user_id, first_name="Ryan", last_name="Parker", gender="Male", date_created=now),
        Guest(user_id=user_id, first_name="Zoe", last_name="Evans", gender="Female", date_created=now),
        Guest(user_id=user_id, first_name="Nathan", last_name="Collins", gender="Male", notes="Photographer", date_created=now),
        Guest(user_id=user_id, first_name="Hannah", last_name="Stewart", gender="Female", date_created=now),
    ]
    db.session.add_all(guests)
    db.session.flush()

    invitations = [
        # New Year's Eve Bash (past) — 12 guests, target 20
        Invitation(event_id=events[0].id, guest_id=guests[0].id, status="Attending", channel="Email", date_invited=date(2025, 11, 20), date_responded=date(2025, 11, 25)),
        Invitation(event_id=events[0].id, guest_id=guests[1].id, status="Attending", channel="WhatsApp", date_invited=date(2025, 11, 20), date_responded=date(2025, 11, 22)),
        Invitation(event_id=events[0].id, guest_id=guests[2].id, status="Attending", channel="Email", date_invited=date(2025, 11, 20), date_responded=date(2025, 12, 1)),
        Invitation(event_id=events[0].id, guest_id=guests[4].id, status="Attending", channel="SMS", date_invited=date(2025, 11, 25), date_responded=date(2025, 11, 28)),
        Invitation(event_id=events[0].id, guest_id=guests[6].id, status="Declined", channel="Email", date_invited=date(2025, 11, 20), date_responded=date(2025, 11, 30), notes="Out of town"),
        Invitation(event_id=events[0].id, guest_id=guests[8].id, status="Attending", channel="Live", date_invited=date(2025, 11, 25), date_responded=date(2025, 11, 26)),
        Invitation(event_id=events[0].id, guest_id=guests[12].id, status="Attending", channel="WhatsApp", date_invited=date(2025, 12, 1), date_responded=date(2025, 12, 5)),
        Invitation(event_id=events[0].id, guest_id=guests[14].id, status="Attending", channel="Email", date_invited=date(2025, 12, 1), date_responded=date(2025, 12, 3)),
        Invitation(event_id=events[0].id, guest_id=guests[16].id, status="Declined", channel="Call", date_invited=date(2025, 12, 1), date_responded=date(2025, 12, 10)),
        Invitation(event_id=events[0].id, guest_id=guests[19].id, status="Attending", channel="Email", date_invited=date(2025, 12, 5), date_responded=date(2025, 12, 8)),
        Invitation(event_id=events[0].id, guest_id=guests[20].id, status="Attending", channel="SMS", date_invited=date(2025, 12, 5), date_responded=date(2025, 12, 9)),
        Invitation(event_id=events[0].id, guest_id=guests[23].id, status="Attending", channel="Email", date_invited=date(2025, 12, 5), date_responded=date(2025, 12, 7), notes="Taking photos"),
        # Q4 Board Meeting (past) — 6 guests
        Invitation(event_id=events[1].id, guest_id=guests[1].id, status="Attending", channel="Email", date_invited=date(2025, 12, 10), date_responded=date(2025, 12, 12)),
        Invitation(event_id=events[1].id, guest_id=guests[5].id, status="Attending", channel="Email", date_invited=date(2025, 12, 10), date_responded=date(2025, 12, 11)),
        Invitation(event_id=events[1].id, guest_id=guests[17].id, status="Attending", channel="Email", date_invited=date(2025, 12, 10), date_responded=date(2025, 12, 13), notes="Presenting Q4 numbers"),
        Invitation(event_id=events[1].id, guest_id=guests[3].id, status="Declined", channel="Email", date_invited=date(2025, 12, 10), date_responded=date(2025, 12, 15)),
        Invitation(event_id=events[1].id, guest_id=guests[9].id, status="Attending", channel="Call", date_invited=date(2025, 12, 12), date_responded=date(2025, 12, 14)),
        Invitation(event_id=events[1].id, guest_id=guests[15].id, status="Pending", channel="Email", date_invited=date(2025, 12, 15)),
        # Valentine's Dinner (past) — 8 guests, target 8
        Invitation(event_id=events[2].id, guest_id=guests[0].id, status="Attending", channel="Email", date_invited=date(2026, 1, 15), date_responded=date(2026, 1, 18)),
        Invitation(event_id=events[2].id, guest_id=guests[1].id, status="Attending", channel="WhatsApp", date_invited=date(2026, 1, 15), date_responded=date(2026, 1, 17)),
        Invitation(event_id=events[2].id, guest_id=guests[2].id, status="Attending", channel="Email", date_invited=date(2026, 1, 15), date_responded=date(2026, 1, 20)),
        Invitation(event_id=events[2].id, guest_id=guests[4].id, status="Attending", channel="Live", date_invited=date(2026, 1, 18), date_responded=date(2026, 1, 19)),
        Invitation(event_id=events[2].id, guest_id=guests[12].id, status="Attending", channel="WhatsApp", date_invited=date(2026, 1, 18), date_responded=date(2026, 1, 22)),
        Invitation(event_id=events[2].id, guest_id=guests[13].id, status="Attending", channel="Email", date_invited=date(2026, 1, 18), date_responded=date(2026, 1, 21)),
        Invitation(event_id=events[2].id, guest_id=guests[18].id, status="Attending", channel="SMS", date_invited=date(2026, 1, 20), date_responded=date(2026, 1, 25)),
        Invitation(event_id=events[2].id, guest_id=guests[20].id, status="Attending", channel="Email", date_invited=date(2026, 1, 20), date_responded=date(2026, 1, 24), notes="Anniversary celebration"),
        # Annual Gala Dinner — 10 guests, target 10
        Invitation(event_id=events[3].id, guest_id=guests[0].id, status="Attending", channel="Email", date_invited=date(2026, 2, 1), date_responded=date(2026, 2, 5), notes="Confirmed plus one"),
        Invitation(event_id=events[3].id, guest_id=guests[1].id, status="Attending", channel="WhatsApp", date_invited=date(2026, 2, 1), date_responded=date(2026, 2, 3)),
        Invitation(event_id=events[3].id, guest_id=guests[2].id, status="Pending", channel="Email", date_invited=date(2026, 2, 1)),
        Invitation(event_id=events[3].id, guest_id=guests[3].id, status="Declined", channel="Call", date_invited=date(2026, 2, 1), date_responded=date(2026, 2, 10), notes="Travel conflict"),
        Invitation(event_id=events[3].id, guest_id=guests[4].id, status="Attending", channel="Email", date_invited=date(2026, 2, 5), date_responded=date(2026, 2, 8)),
        Invitation(event_id=events[3].id, guest_id=guests[8].id, status="Attending", channel="Live", date_invited=date(2026, 2, 5), date_responded=date(2026, 2, 7)),
        Invitation(event_id=events[3].id, guest_id=guests[10].id, status="Not Sent"),
        Invitation(event_id=events[3].id, guest_id=guests[11].id, status="Attending", channel="Email", date_invited=date(2026, 2, 1), date_responded=date(2026, 2, 2), notes="VIP table"),
        Invitation(event_id=events[3].id, guest_id=guests[14].id, status="Attending", channel="WhatsApp", date_invited=date(2026, 2, 8), date_responded=date(2026, 2, 12)),
        Invitation(event_id=events[3].id, guest_id=guests[21].id, status="Pending", channel="Email", date_invited=date(2026, 2, 10)),
        # Team Building Retreat — 8 guests, target 6
        Invitation(event_id=events[4].id, guest_id=guests[1].id, status="Attending", channel="WhatsApp", date_invited=date(2026, 3, 1), date_responded=date(2026, 3, 5)),
        Invitation(event_id=events[4].id, guest_id=guests[3].id, status="Attending", channel="Email", date_invited=date(2026, 3, 1), date_responded=date(2026, 3, 4)),
        Invitation(event_id=events[4].id, guest_id=guests[5].id, status="Pending", channel="Call", date_invited=date(2026, 3, 1), notes="Call back tomorrow"),
        Invitation(event_id=events[4].id, guest_id=guests[7].id, status="Declined", channel="WhatsApp", date_invited=date(2026, 3, 1), date_responded=date(2026, 3, 8)),
        Invitation(event_id=events[4].id, guest_id=guests[9].id, status="Not Sent"),
        Invitation(event_id=events[4].id, guest_id=guests[11].id, status="Not Sent"),
        Invitation(event_id=events[4].id, guest_id=guests[15].id, status="Attending", channel="Email", date_invited=date(2026, 3, 5), date_responded=date(2026, 3, 7)),
        Invitation(event_id=events[4].id, guest_id=guests[19].id, status="Attending", channel="SMS", date_invited=date(2026, 3, 5), date_responded=date(2026, 3, 9)),
        # Summer Garden Party — 7 guests, no target
        Invitation(event_id=events[5].id, guest_id=guests[0].id, status="Attending", channel="Live", date_invited=date(2026, 4, 1), date_responded=date(2026, 4, 3)),
        Invitation(event_id=events[5].id, guest_id=guests[2].id, status="Attending", channel="WhatsApp", date_invited=date(2026, 4, 1), date_responded=date(2026, 4, 2)),
        Invitation(event_id=events[5].id, guest_id=guests[4].id, status="Pending", channel="Email", date_invited=date(2026, 4, 5)),
        Invitation(event_id=events[5].id, guest_id=guests[6].id, status="Attending", channel="SMS", date_invited=date(2026, 4, 1), date_responded=date(2026, 4, 4)),
        Invitation(event_id=events[5].id, guest_id=guests[8].id, status="Declined", channel="Call", date_invited=date(2026, 4, 1), date_responded=date(2026, 4, 7), notes="Health reasons"),
        Invitation(event_id=events[5].id, guest_id=guests[16].id, status="Attending", channel="Email", date_invited=date(2026, 4, 8), date_responded=date(2026, 4, 10)),
        Invitation(event_id=events[5].id, guest_id=guests[22].id, status="Not Sent"),
        # Lake House Weekend — 14 guests, target 15
        Invitation(event_id=events[6].id, guest_id=guests[0].id, status="Attending", channel="Email", date_invited=date(2026, 5, 1), date_responded=date(2026, 5, 5)),
        Invitation(event_id=events[6].id, guest_id=guests[2].id, status="Attending", channel="Email", date_invited=date(2026, 5, 1), date_responded=date(2026, 5, 3)),
        Invitation(event_id=events[6].id, guest_id=guests[4].id, status="Attending", channel="WhatsApp", date_invited=date(2026, 5, 1), date_responded=date(2026, 5, 4)),
        Invitation(event_id=events[6].id, guest_id=guests[6].id, status="Pending", channel="Email", date_invited=date(2026, 5, 1)),
        Invitation(event_id=events[6].id, guest_id=guests[8].id, status="Attending", channel="Live", date_invited=date(2026, 5, 5), date_responded=date(2026, 5, 6)),
        Invitation(event_id=events[6].id, guest_id=guests[10].id, status="Attending", channel="Email", date_invited=date(2026, 5, 5), date_responded=date(2026, 5, 9)),
        Invitation(event_id=events[6].id, guest_id=guests[12].id, status="Declined", channel="Call", date_invited=date(2026, 5, 5), date_responded=date(2026, 5, 12), notes="Scheduling conflict"),
        Invitation(event_id=events[6].id, guest_id=guests[14].id, status="Attending", channel="Email", date_invited=date(2026, 5, 8), date_responded=date(2026, 5, 11)),
        Invitation(event_id=events[6].id, guest_id=guests[16].id, status="Attending", channel="WhatsApp", date_invited=date(2026, 5, 8), date_responded=date(2026, 5, 10)),
        Invitation(event_id=events[6].id, guest_id=guests[18].id, status="Not Sent"),
        Invitation(event_id=events[6].id, guest_id=guests[20].id, status="Attending", channel="SMS", date_invited=date(2026, 5, 10), date_responded=date(2026, 5, 14), notes="Bringing plus one"),
        Invitation(event_id=events[6].id, guest_id=guests[22].id, status="Pending", channel="Email", date_invited=date(2026, 5, 10)),
        Invitation(event_id=events[6].id, guest_id=guests[23].id, status="Attending", channel="Email", date_invited=date(2026, 5, 1), date_responded=date(2026, 5, 2), notes="Bringing kayak"),
        Invitation(event_id=events[6].id, guest_id=guests[24].id, status="Attending", channel="WhatsApp", date_invited=date(2026, 5, 10), date_responded=date(2026, 5, 13)),
        # Fall Pheasant Hunt — 6 guests, target 4
        Invitation(event_id=events[7].id, guest_id=guests[5].id, status="Not Sent"),
        Invitation(event_id=events[7].id, guest_id=guests[7].id, status="Not Sent"),
        Invitation(event_id=events[7].id, guest_id=guests[9].id, status="Attending", channel="Email", date_invited=date(2026, 5, 1), date_responded=date(2026, 5, 3)),
        Invitation(event_id=events[7].id, guest_id=guests[11].id, status="Attending", channel="Email", date_invited=date(2026, 5, 1), date_responded=date(2026, 5, 2), notes="Bringing own gear"),
        Invitation(event_id=events[7].id, guest_id=guests[17].id, status="Attending", channel="Call", date_invited=date(2026, 5, 5), date_responded=date(2026, 5, 8)),
        Invitation(event_id=events[7].id, guest_id=guests[23].id, status="Pending", channel="Email", date_invited=date(2026, 5, 10)),
    ]
    db.session.add_all(invitations)
    db.session.commit()


# ── Run ───────────────────────────────────────────────────────────────────────

with app.app_context():
    db.create_all()
    # Migrate existing databases: add new columns if missing
    # Each statement runs in its own SAVEPOINT so a failure doesn't abort the transaction (PostgreSQL)
    with db.engine.connect() as conn:
        for stmt in [
            "ALTER TABLE invitation ADD COLUMN channel VARCHAR(50) DEFAULT ''",
            "ALTER TABLE invitation ADD COLUMN notes TEXT DEFAULT ''",
            "ALTER TABLE guest ADD COLUMN is_me BOOLEAN DEFAULT 0",
            "ALTER TABLE event ADD COLUMN user_id INTEGER",
            "ALTER TABLE guest ADD COLUMN user_id INTEGER",
            "ALTER TABLE event ADD COLUMN date_created DATE",
            "ALTER TABLE guest ADD COLUMN date_created TIMESTAMP",
            "ALTER TABLE guest ADD COLUMN date_edited TIMESTAMP",
            "ALTER TABLE event ADD COLUMN date_edited TIMESTAMP",
            "ALTER TABLE event ADD COLUMN target_attendees INTEGER",
        ]:
            try:
                conn.execute(db.text("SAVEPOINT sp"))
                conn.execute(db.text(stmt))
                conn.execute(db.text("RELEASE SAVEPOINT sp"))
            except Exception:
                conn.execute(db.text("ROLLBACK TO SAVEPOINT sp"))
        conn.commit()
if __name__ == "__main__":
    app.run(debug=os.environ.get("FLASK_DEBUG", "0") == "1")
