import re
from io import BytesIO
from flask import send_file, make_response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")


def _styled_sheet(wb, title, headers):
    ws = wb.active
    ws.title = re.sub(r"[/\\*?\[\]:]", "_", title)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    return ws


def _to_download(wb, filename):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, download_name=filename, as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def export_events_xlsx(events):
    wb = Workbook()
    ws = _styled_sheet(wb, "Events", ["Name", "Type", "Date", "Location", "Invited", "Attending", "Notes"])
    for e in events:
        attending = sum(1 for inv in e.invitations if inv.status == "Attending")
        ws.append([e.name, e.event_type, e.date.strftime("%Y-%m-%d"), e.location or "",
                   len(e.invitations), attending, e.notes or ""])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    return _to_download(wb, "events.xlsx")


def export_guests_xlsx(guests):
    wb = Workbook()
    ws = _styled_sheet(wb, "Guests", ["Last Name", "First Name", "Gender", "Notes"])
    for g in guests:
        ws.append([g.last_name or "", g.first_name, g.gender, g.notes or ""])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    return _to_download(wb, "guests.xlsx")


def _get_seating_info(inv):
    """Get table name and seat number separately for an invitation."""
    assignments = inv.seat_assignment
    if not assignments:
        return "", ""
    sa = assignments[0] if isinstance(assignments, list) else assignments
    table = sa.table
    label = table.label or ("Table " + str(table.table_number))
    return label, str(sa.seat_position)


def export_event_guests_xlsx(event):
    wb = Workbook()
    ws = _styled_sheet(wb, event.name[:31],
                       ["Last Name", "First Name", "Gender", "Tags", "Sent",
                        "Invited On", "Status", "Responded On", "Table", "Seat", "Inv. Notes", "Guest Notes"])
    for inv in event.invitations:
        g = inv.guest
        if g.deleted_at:
            continue
        sent = "Yes" if inv.status != "Not Sent" else "No"
        tags = ", ".join(t.name for t in g.tags if not t.deleted_at)
        table_name, seat_num = _get_seating_info(inv)
        ws.append([g.last_name or "", g.first_name, g.gender, tags, sent,
                   inv.date_invited.strftime("%Y-%m-%d") if inv.date_invited else "",
                   inv.status,
                   inv.date_responded.strftime("%Y-%m-%d") if inv.date_responded else "",
                   table_name, seat_num,
                   inv.notes or "", g.notes or ""])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    safe_name = re.sub(r"[^\w\-]", "_", event.name).strip("_").lower()
    date_str = event.date.strftime("%Y-%m-%d") if event.date else ""
    return _to_download(wb, f"{safe_name}_{date_str}_guests.xlsx")


def export_event_guests_text(event):
    """Export attending/pending guests as formatted text for sharing."""
    attending = []
    pending = []
    for inv in event.invitations:
        if inv.guest.deleted_at:
            continue
        name = inv.guest.full_name
        if inv.status == "Attending":
            attending.append(name)
        elif inv.status == "Pending":
            pending.append(name)

    attending.sort(key=str.lower)
    pending.sort(key=str.lower)

    lines = []
    lines.append(f"{event.name} ({event.date.strftime('%d %B %Y')})")
    summary = f"✅ {len(attending)} attending"
    if len(pending) > 0:
        summary += f" · 🟠 {len(pending)} pending"
    lines.append(summary)
    lines.append("")

    if attending:
        lines.append("— Attending —")
        for name in attending:
            lines.append(f"• {name}")
        lines.append("")

    if pending:
        lines.append("— Pending —")
        for name in pending:
            lines.append(f"• {name}")
        lines.append("")

    text = "\n".join(lines)
    response = make_response(text)
    response.headers["Content-Type"] = "text/plain; charset=utf-8"
    return response
